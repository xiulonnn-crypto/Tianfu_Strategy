#!/usr/bin/env python3
"""
从天府计划回测 Excel 导出 data/backtest/v1.3.1-{10y|20y|30y}-{summary|nav|trades}.json

用法:
  python3 scripts/import_backtest.py \\
    --10y /path/to/天府计划1.3.1回测10年.xlsx \\
    --20y /path/to/天府计划1.3.1回测20年.xlsx \\
    --30y /path/to/天府计划1.3.1回测30年.xlsx

  python3 scripts/import_backtest.py --dry-run ...  # 只打印统计，不写文件
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable

try:
    import openpyxl
except ImportError:
    print("需要: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


def _parse_pct(s: Any) -> float | None:
    if s is None:
        return None
    t = str(s).strip().replace("%", "").replace(",", "")
    if not t or t.lower() == "none":
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _parse_money(s: Any) -> float | None:
    if s is None:
        return None
    t = str(s).strip().replace("$", "").replace(",", "").replace("￥", "")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _parse_int_days(s: Any) -> int | None:
    if s is None:
        return None
    t = str(s).strip().replace("天", "").replace(",", "")
    if not t:
        return None
    try:
        return int(float(t))
    except ValueError:
        return None


def _parse_symbols(cell: Any) -> list[str]:
    if cell is None:
        return []
    raw = str(cell).strip()
    parts = re.split(r"[,，]", raw)
    return [p.strip() for p in parts if p.strip()]


def parse_summary_sheet(ws) -> dict[str, Any]:
    """指标汇总 sheet -> meta + metrics dict."""
    cells: dict[str, Any] = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        k = str(row[0]).strip()
        cells[k] = row[1] if len(row) > 1 else None

    symbols = _parse_symbols(cells.get("标的"))
    start = cells.get("起始日期")
    end = cells.get("结束日期")
    comm_raw = cells.get("手续费率")
    comm_rate = None
    if comm_raw is not None:
        m = re.search(r"([\d.]+)\s*%", str(comm_raw))
        if m:
            comm_rate = float(m.group(1)) / 100.0
        else:
            comm_rate = _parse_pct(comm_raw)
            if comm_rate is not None and comm_rate > 0.5:
                comm_rate = comm_rate / 100.0

    slip = cells.get("滑点(bps)")
    try:
        slip_bps = int(float(slip)) if slip is not None else None
    except (TypeError, ValueError):
        slip_bps = None

    metrics = {
        "cumulative_return_pct": _parse_pct(cells.get("累积收益")),
        "cagr_pct": _parse_pct(cells.get("年化收益")),
        "max_drawdown_pct": _parse_pct(cells.get("最大回撤")),
        "recovery_days": _parse_int_days(cells.get("回撤恢复天数")),
        "sharpe": _parse_pct(cells.get("夏普比率")),  # stored as number in Excel
        "win_rate_pct": _parse_pct(cells.get("胜率")),
        "profit_loss_ratio": _parse_pct(cells.get("盈亏比")),
        "trade_count": int(float(str(cells.get("交易次数", 0)).replace(",", "")))
        if cells.get("交易次数") is not None
        else None,
        "final_capital": _parse_money(cells.get("最终资金")),
        "alpha_pct": _parse_pct(cells.get("Alpha")),
        "beta": _parse_pct(cells.get("Beta")),
    }
    # 夏普可能是 0.157 这种，_parse_pct works
    if metrics["sharpe"] is not None and abs(metrics["sharpe"]) > 10:
        # mis-parsed as percent
        try:
            metrics["sharpe"] = float(str(cells.get("夏普比率")).replace(",", ""))
        except (TypeError, ValueError):
            pass

    init_cap = _parse_money(cells.get("初始资金"))
    meta = {
        "initial_capital": init_cap,
        "commission_rate": comm_rate,
        "slippage_bps": slip_bps,
        "symbols": symbols,
        "start_date": str(start)[:10] if start else None,
        "end_date": str(end)[:10] if end else None,
    }
    return {"meta": meta, "metrics": metrics}


def parse_nav_sheet(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        date, nav, cash, hold, dd = (list(row) + [None] * 5)[:5]
        if date is None:
            continue
        dstr = str(date)[:10] if not isinstance(date, str) else str(date)[:10]
        dd_pct = _parse_pct(dd)
        rows.append(
            {
                "date": dstr,
                "nav": float(nav) if nav is not None else None,
                "cash": float(cash) if cash is not None else None,
                "holdings_value": float(hold) if hold is not None else None,
                "drawdown_pct": dd_pct,
            }
        )
    return rows


def parse_trades_sheet(ws) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        seq, date, side, sym, price, qty, fee, slip, pnl = (list(row) + [None] * 9)[:9]
        if date is None:
            continue
        dstr = str(date)[:10] if not isinstance(date, str) else str(date)[:10]
        side_s = str(side).strip().upper() if side else ""
        if side_s in ("BUY", "买入"):
            side_norm = "BUY"
        elif side_s in ("SELL", "卖出"):
            side_norm = "SELL"
        else:
            side_norm = side_s
        out.append(
            {
                "seq": int(seq) if seq is not None else len(out) + 1,
                "date": dstr,
                "side": side_norm,
                "symbol": str(sym).strip() if sym else "",
                "price": float(price) if price is not None else None,
                "qty": float(qty) if qty is not None else None,
                "commission": float(fee) if fee is not None else None,
                "slippage": float(slip) if slip is not None else None,
                "pnl": float(pnl) if pnl is not None else None,
            }
        )
    return out


def _parse_date(d: str):
    return datetime.strptime(d[:10], "%Y-%m-%d").date()


BENCHMARK_SYMBOL = "^IXIC"  # 纳斯达克综合指数，用于 CAPM 回归
QQQ_SYMBOL = "QQQ"
QQQ_IPO_DATE = "1999-03-10"


def _fetch_daily_closes(symbol: str, start_date: str, end_date: str) -> dict[str, float]:
    """逐标的拉取日线收盘，返回 {date_str: close}。"""
    try:
        import yfinance as yf
    except ImportError as e:
        raise RuntimeError("需要 yfinance：pip3 install yfinance") from e

    start_dt = _parse_date(start_date)
    end_dt = _parse_date(end_date)
    hist = yf.Ticker(symbol).history(
        start=start_dt.isoformat(),
        end=(end_dt + timedelta(days=3)).isoformat(),
        auto_adjust=True,
    )
    out: dict[str, float] = {}
    for idx, row in hist.iterrows():
        d = str(idx)[:10]
        try:
            out[d] = float(row["Close"])
        except (KeyError, TypeError, ValueError):
            continue
    return out


def _align_close_series(nav_dates: list[str], closes: dict[str, float]) -> list[float | None]:
    """按 nav 日期顺序，缺失日用上一有效收盘（前向填充）。"""
    last: float | None = None
    out: list[float | None] = []
    for d in nav_dates:
        if d in closes:
            last = closes[d]
        out.append(last)
    return out


def _build_qqq_proxy_series(
    nav_dates: list[str],
    qqq_closes: dict[str, float],
    ixic_closes: dict[str, float],
) -> tuple[list[float | None], int, str | None]:
    """
    构造连续 QQQ proxy：QQQ IPO 前用 ^IXIC × k 缩放衔接。
    返回 (px 与 nav_dates 等长, proxy_days, first_qqq_date)。
    """
    if not nav_dates:
        return [], 0, None

    first_qqq_date: str | None = None
    for d in nav_dates:
        if d in qqq_closes:
            first_qqq_date = d
            break
    if first_qqq_date is None:
        return [None] * len(nav_dates), 0, None

    qqq_aligned = _align_close_series(nav_dates, qqq_closes)
    ixic_aligned = _align_close_series(nav_dates, ixic_closes)

    i0 = nav_dates.index(first_qqq_date)
    q0 = qqq_aligned[i0]
    i0_ix = ixic_aligned[i0]
    if q0 is None or i0_ix is None or i0_ix <= 0:
        return [None] * len(nav_dates), 0, None

    k = q0 / i0_ix
    proxy_days = sum(1 for d in nav_dates if d < first_qqq_date)

    px: list[float | None] = []
    for i, d in enumerate(nav_dates):
        if d >= first_qqq_date:
            v = qqq_aligned[i]
            px.append(v)
        else:
            ix = ixic_aligned[i]
            px.append(k * ix if ix is not None and ix > 0 else None)

    return px, proxy_days, first_qqq_date


def _group_trades_net_cash_flow_by_date(trades: list[dict[str, Any]]) -> dict[str, float]:
    """每日交易导致的现金净流入（卖出为正，买入为负）。含手续费与滑点。"""
    by: dict[str, float] = {}
    for t in trades:
        d = str(t.get("date", ""))[:10]
        if not d:
            continue
        side = str(t.get("side", "")).upper()
        price = float(t["price"]) if t.get("price") is not None else 0.0
        qty = float(t["qty"]) if t.get("qty") is not None else 0.0
        comm = float(t["commission"]) if t.get("commission") is not None else 0.0
        slip = float(t["slippage"]) if t.get("slippage") is not None else 0.0
        if side == "BUY":
            flow = -(price * qty + comm + slip)
        elif side == "SELL":
            flow = price * qty - comm - slip
        else:
            continue
        by[d] = by.get(d, 0.0) + flow
    return by


def compute_port_twr_pct_series(nav_rows: list[dict[str, Any]], trades: list[dict[str, Any]]) -> list[float | None]:
    """
    日链式 TWR（inject = Δcash − 交易流，含滑点）。用于测试或对照；
    实盘回测含分红/利息等未在 trades 中体现时，与 Excel cumulative_return 易偏离。
    """
    if not nav_rows:
        return []
    trade_flow = _group_trades_net_cash_flow_by_date(trades)
    out: list[float | None] = []
    cum = 1.0
    prev_nav: float | None = None
    prev_cash: float | None = None

    for row in nav_rows:
        nav = row.get("nav")
        cash = row.get("cash")
        if nav is None or cash is None:
            out.append(round((cum - 1.0) * 100.0, 4))
            continue
        try:
            nav_f = float(nav)
            cash_f = float(cash)
        except (TypeError, ValueError):
            out.append(round((cum - 1.0) * 100.0, 4))
            continue

        if prev_nav is None or prev_cash is None:
            out.append(0.0)
            prev_nav, prev_cash = nav_f, cash_f
            continue

        d = str(row.get("date", ""))[:10]
        flow = trade_flow.get(d, 0.0)
        inject = (cash_f - prev_cash) - flow
        denom = prev_nav + inject
        if denom <= 0:
            r = 0.0
        else:
            r = nav_f / denom - 1.0
        cum *= 1.0 + r
        out.append(round((cum - 1.0) * 100.0, 4))
        prev_nav, prev_cash = nav_f, cash_f

    return out


def compute_port_return_pct_aligned(
    nav_rows: list[dict[str, Any]],
    cumulative_return_pct: float | None,
) -> list[float | None]:
    """
    组合累计收益曲线：按净值相对首日的比例形状，线性缩放到与 summary 的
    cumulative_return_pct 末值一致（Excel 口径含分红等，无法仅用 trades 复现日度 TWR）。
    """
    if not nav_rows:
        return []
    n0 = nav_rows[0].get("nav")
    if n0 is None:
        return [None] * len(nav_rows)
    try:
        n0f = float(n0)
    except (TypeError, ValueError):
        return [None] * len(nav_rows)
    if n0f <= 0:
        return [None] * len(nav_rows)

    n_last = nav_rows[-1].get("nav")
    try:
        raw_end = float(n_last) / n0f - 1.0 if n_last is not None else 0.0
    except (TypeError, ValueError):
        raw_end = 0.0

    if cumulative_return_pct is None:
        tgt_ratio = raw_end
    else:
        tgt_ratio = float(cumulative_return_pct) / 100.0

    scale = (tgt_ratio / raw_end) if abs(raw_end) > 1e-12 else 1.0

    out: list[float | None] = []
    for row in nav_rows:
        nav = row.get("nav")
        if nav is None:
            out.append(None)
            continue
        try:
            raw = float(nav) / n0f - 1.0
        except (TypeError, ValueError):
            out.append(None)
            continue
        out.append(round(raw * scale * 100.0, 4))

    return out


def compute_qqq_bh_pct_series(px: list[float | None]) -> list[float | None]:
    if not px:
        return []
    base = None
    for v in px:
        if v is not None and v > 0:
            base = v
            break
    if base is None or base <= 0:
        return [None] * len(px)
    out: list[float | None] = []
    for v in px:
        if v is None or v <= 0:
            out.append(None)
        else:
            out.append(round((v / base - 1.0) * 100.0, 4))
    return out


def _calendar_months_span(first_date: str, last_date: str) -> int:
    a = _parse_date(first_date)
    b = _parse_date(last_date)
    return (b.year - a.year) * 12 + (b.month - a.month) + 1


def compute_qqq_dca_pct_series(
    nav_dates: list[str],
    px: list[float | None],
    initial_capital: float | None,
) -> list[float | None]:
    """每月第一个有 px 的交易日定投，总额 initial_capital 均摊到日历月数。"""
    n = len(nav_dates)
    if n == 0 or initial_capital is None or initial_capital <= 0:
        return [None] * n

    months_n = _calendar_months_span(nav_dates[0], nav_dates[-1])
    if months_n <= 0:
        return [None] * n

    monthly = initial_capital / float(months_n)
    cum_shares = 0.0
    cum_cost = 0.0
    invested_months: set[tuple[int, int]] = set()

    # 预计算：每月首个有效 px 的索引
    month_first_idx: dict[tuple[int, int], int] = {}
    for i, d in enumerate(nav_dates):
        if px[i] is None or px[i] <= 0:
            continue
        dt = _parse_date(d)
        ym = (dt.year, dt.month)
        if ym not in month_first_idx:
            month_first_idx[ym] = i

    out: list[float | None] = []
    for i, d in enumerate(nav_dates):
        dt = _parse_date(d)
        ym = (dt.year, dt.month)
        if ym not in invested_months and month_first_idx.get(ym) == i:
            pr = px[i]
            if pr is not None and pr > 0:
                cum_shares += monthly / pr
                cum_cost += monthly
                invested_months.add(ym)

        price = px[i]
        if cum_cost > 0 and price is not None and price > 0:
            out.append(round((cum_shares * price / cum_cost - 1.0) * 100.0, 4))
        else:
            out.append(None)

    return out


def enrich_benchmark(
    out_dir: Path,
    version: str,
    periods: list[str],
    force: bool = False,
    fetch_closes: Callable[[str, str, str], dict[str, float]] | None = None,
) -> None:
    """
    为 nav.json 写入 port_ret_pct / qqq_bh_pct / qqq_dca_pct，并更新 summary.benchmark。
    fetch_closes 可注入以便测试 mock。
    """
    fetch = fetch_closes or _fetch_daily_closes

    for period in periods:
        nav_path = out_dir / f"{version}-{period}-nav.json"
        trades_path = out_dir / f"{version}-{period}-trades.json"
        summary_path = out_dir / f"{version}-{period}-summary.json"
        if not nav_path.is_file() or not summary_path.is_file():
            print(f"[skip] {period}: 缺少 nav 或 summary", file=sys.stderr)
            continue

        nav_rows: list[dict[str, Any]] = json.loads(nav_path.read_text(encoding="utf-8"))
        trades: list[dict[str, Any]] = []
        if trades_path.is_file():
            trades = json.loads(trades_path.read_text(encoding="utf-8"))

        if not force and nav_rows and "port_ret_pct" in nav_rows[0]:
            print(f"[{period}] 跳过（已有 enrich 字段）；加 --force 覆盖")
            continue

        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        start = summary.get("start_date") or nav_rows[0].get("date")
        end = summary.get("end_date") or nav_rows[-1].get("date")
        if not start or not end:
            print(f"[skip] {period}: 无起止日期", file=sys.stderr)
            continue

        initial_capital = summary.get("initial_capital")
        if isinstance(initial_capital, (int, float)):
            ic = float(initial_capital)
        else:
            ic = None

        nav_dates = [str(r.get("date", ""))[:10] for r in nav_rows]
        nav_min = min(nav_dates) if nav_dates else start
        try:
            qqq = fetch(QQQ_SYMBOL, start, end)
            if nav_min < QQQ_IPO_DATE:
                ixic = fetch(BENCHMARK_SYMBOL, start, end)
            else:
                ixic = {}
        except Exception as e:  # noqa: BLE001
            print(f"[warn] {period}: 拉取行情失败，跳过：{e}", file=sys.stderr)
            continue

        px, proxy_days, first_qqq = _build_qqq_proxy_series(nav_dates, qqq, ixic)

        m = summary.get("metrics") or {}
        cum_pct = m.get("cumulative_return_pct")
        cum_pct_f = float(cum_pct) if isinstance(cum_pct, (int, float)) else None

        port_pct = compute_port_return_pct_aligned(nav_rows, cum_pct_f)
        bh_pct = compute_qqq_bh_pct_series(px)
        dca_pct = compute_qqq_dca_pct_series(nav_dates, px, ic)

        for i, row in enumerate(nav_rows):
            row["port_ret_pct"] = port_pct[i] if i < len(port_pct) else None
            row["qqq_bh_pct"] = bh_pct[i] if i < len(bh_pct) else None
            row["qqq_dca_pct"] = dca_pct[i] if i < len(dca_pct) else None

        summary["benchmark"] = {
            "symbol": QQQ_SYMBOL,
            "proxy_before": BENCHMARK_SYMBOL,
            "qqq_ipo_date": QQQ_IPO_DATE,
            "proxy_days": proxy_days,
            "dca_total": ic,
            "dca_schedule": "monthly",
            "first_qqq_date": first_qqq,
            "portfolio_curve": "nav_scaled_to_summary_cumulative_pct",
        }

        nav_path.write_text(
            json.dumps(nav_rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        summary_path.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        print(f"[{period}] enrich_benchmark 已写入（proxy_days={proxy_days}）")


def _winsorize(values: list[float], lo_pct: float = 1.0, hi_pct: float = 99.0) -> list[float]:
    """按百分位夹紧极值。用于剔除资金注入日引起的异常日收益尖峰。"""
    if not values:
        return []
    s = sorted(values)
    n = len(s)
    lo_i = max(0, min(n - 1, int(n * lo_pct / 100.0)))
    hi_i = max(0, min(n - 1, int(n * hi_pct / 100.0)))
    lo_v = s[lo_i]
    hi_v = s[hi_i]
    return [max(lo_v, min(hi_v, v)) for v in values]


def compute_alpha_beta_from_returns(
    port_returns: list[float],
    bench_returns: list[float],
    cagr_port_annual: float | None = None,
) -> tuple[float | None, float | None]:
    """
    根据对齐后的组合 / 基准日收益率序列，按 CAPM 口径计算 Jensen's Alpha（年化）与 Beta。

    β  = Cov(Rp_winsorized, Rb) / Var(Rb)
    α_年化% = (CAGR_port - β × CAGR_bench) × 100

    参数:
      port_returns, bench_returns: 同日对齐的等长日收益序列
      cagr_port_annual: 若提供（如来自 summary.metrics.cagr_pct/100），
                        作为 CAGR_port 使用；否则从 port_returns 推算。
                        回测 NAV 若含周期性资金注入，直接推算会显著偏高，
                        应传入 summary 的可信 CAGR。

    返回 (alpha_pct, beta)，数据不足时返回 (None, None)。
    """
    if not port_returns or not bench_returns:
        return (None, None)
    if len(port_returns) != len(bench_returns):
        return (None, None)
    n = len(port_returns)
    if n < 2:
        return (None, None)

    rp_w = _winsorize(port_returns)
    mean_p = sum(rp_w) / n
    mean_b = sum(bench_returns) / n
    cov_pb = sum((p - mean_p) * (b - mean_b) for p, b in zip(rp_w, bench_returns)) / n
    var_b = sum((b - mean_b) ** 2 for b in bench_returns) / n
    if var_b < 1e-18:
        return (None, None)
    beta = cov_pb / var_b

    years = n / 252.0
    if years <= 0:
        return (None, None)

    if cagr_port_annual is None:
        prod_p = 1.0
        for p in port_returns:
            prod_p *= 1.0 + p
        cagr_p = prod_p ** (1.0 / years) - 1.0 if prod_p > 0 else -1.0
    else:
        cagr_p = cagr_port_annual

    prod_b = 1.0
    for b in bench_returns:
        prod_b *= 1.0 + b
    cagr_b = prod_b ** (1.0 / years) - 1.0 if prod_b > 0 else -1.0

    alpha_pct = (cagr_p - beta * cagr_b) * 100.0
    return (round(alpha_pct, 2), round(beta, 3))


def _nav_dedup_sort(nav_rows: list[dict[str, Any]]) -> list[tuple[str, float]]:
    """nav_rows → [(date, nav), ...]，按日期去重（保留最后一次）并升序。"""
    by_date: dict[str, float] = {}
    for r in nav_rows:
        d = r.get("date")
        v = r.get("nav")
        if d is None or v is None:
            continue
        try:
            by_date[str(d)[:10]] = float(v)
        except (TypeError, ValueError):
            continue
    return sorted(by_date.items(), key=lambda x: x[0])


def _fetch_benchmark_closes(start_date: str, end_date: str) -> dict[str, float]:
    """逐标的循环拉取 ^IXIC 日线收盘，返回 {date_str: close}。"""
    return _fetch_daily_closes(BENCHMARK_SYMBOL, start_date, end_date)


def compute_alpha_beta_for_nav(
    nav_rows: list[dict[str, Any]],
    cagr_port_annual: float | None = None,
) -> tuple[float | None, float | None]:
    """
    从 nav_rows（未必去重 / 排序）拉取 ^IXIC 基准，对齐到 nav 日期后计算 CAPM α/β。

    cagr_port_annual: 可传入 summary.metrics.cagr_pct/100 作为可信年化，
                      避免因 NAV 含资金注入导致 α 被高估。
    """
    pairs = _nav_dedup_sort(nav_rows)
    if len(pairs) < 3:
        return (None, None)

    start = pairs[0][0]
    end = pairs[-1][0]
    bench = _fetch_benchmark_closes(start, end)
    if not bench:
        return (None, None)

    port_ret: list[float] = []
    bench_ret: list[float] = []
    prev_nav: float | None = None
    prev_bench: float | None = None
    for d, nav in pairs:
        bc = bench.get(d)
        if bc is None:
            # 基准该日无数据（假日错位），跳过以保持同日对齐
            continue
        if prev_nav is not None and prev_bench is not None and prev_nav > 0 and prev_bench > 0:
            port_ret.append(nav / prev_nav - 1.0)
            bench_ret.append(bc / prev_bench - 1.0)
        prev_nav = nav
        prev_bench = bc

    return compute_alpha_beta_from_returns(port_ret, bench_ret, cagr_port_annual)


def recompute_summary_risk(
    out_dir: Path,
    version: str,
    periods: list[str],
    force: bool = False,
) -> None:
    """
    重算现有 summary.json 中的 alpha_pct / beta，不重新解析 Excel。
    使用 data/backtest 目录下的 nav.json + yfinance ^IXIC。

    默认仅在 Excel 值为 0/0「哨兵」时重算（Excel 公式失败的典型表现），
    其它已有非零值保留 Excel 口径。传入 force=True 强制覆盖所有周期。
    """
    for period in periods:
        summary_path = out_dir / f"{version}-{period}-summary.json"
        nav_path = out_dir / f"{version}-{period}-nav.json"
        if not summary_path.is_file() or not nav_path.is_file():
            print(f"[skip] {period}: 缺少 {summary_path.name} 或 {nav_path.name}")
            continue
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        metrics = summary.setdefault("metrics", {})
        old_a = metrics.get("alpha_pct")
        old_b = metrics.get("beta")

        needs_recompute = force or (
            (old_a in (None, 0, 0.0)) and (old_b in (None, 0, 0.0))
        )
        if not needs_recompute:
            print(f"[{period}] 跳过（Excel 值非 0：alpha={old_a}, beta={old_b}）；"
                  f"如需强制覆盖请加 --force")
            continue

        nav_rows = json.loads(nav_path.read_text(encoding="utf-8"))
        cagr_pct = metrics.get("cagr_pct")
        cagr_ann = cagr_pct / 100.0 if isinstance(cagr_pct, (int, float)) else None
        alpha_pct, beta = compute_alpha_beta_for_nav(nav_rows, cagr_ann)
        metrics["alpha_pct"] = alpha_pct
        metrics["beta"] = beta
        summary_path.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        print(f"[{period}] alpha {old_a} → {alpha_pct}  |  beta {old_b} → {beta}")


def compute_top_drawdowns(dates: list[str], nav: list[float], top_n: int = 3) -> list[dict[str, Any]]:
    """自峰值至收复峰值（或序列结束）的回撤事件，按深度降序取前 N。"""
    if len(dates) != len(nav) or len(nav) < 2:
        return []
    episodes: list[dict[str, Any]] = []
    peak_i = 0
    i = 1
    n = len(nav)
    while i < n:
        if nav[i] >= nav[peak_i]:
            peak_i = i
            i += 1
            continue
        trough_i = i
        trough_v = nav[i]
        j = i + 1
        while j < n and nav[j] < nav[peak_i]:
            if nav[j] < trough_v:
                trough_i = j
                trough_v = nav[j]
            j += 1
        depth_pct = round((1.0 - trough_v / nav[peak_i]) * 100.0, 2)
        rec_i = j if j < n else None
        recovery_date = dates[rec_i] if rec_i is not None else None
        pd_peak = _parse_date(dates[peak_i])
        pd_trough = _parse_date(dates[trough_i])
        duration_days = (pd_trough - pd_peak).days
        recovery_days = None
        if rec_i is not None:
            recovery_days = (_parse_date(dates[rec_i]) - pd_trough).days
        episodes.append(
            {
                "peak_date": dates[peak_i],
                "trough_date": dates[trough_i],
                "recovery_date": recovery_date,
                "drawdown_pct": depth_pct,
                "duration_days": duration_days,
                "recovery_days": recovery_days,
            }
        )
        if rec_i is None:
            break
        peak_i = rec_i
        i = rec_i + 1

    episodes.sort(key=lambda e: e["drawdown_pct"], reverse=True)
    return episodes[:top_n]


def process_one_workbook(path: Path, period: str, version: str, dry_run: bool, out_dir: Path) -> None:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "指标汇总" not in wb.sheetnames:
        raise ValueError(f"{path}: 缺少「指标汇总」")
    summary_part = parse_summary_sheet(wb["指标汇总"])
    nav_rows = parse_nav_sheet(wb["每日净值"]) if "每日净值" in wb.sheetnames else []
    trades = parse_trades_sheet(wb["交易明细"]) if "交易明细" in wb.sheetnames else []

    dates = [r["date"] for r in nav_rows if r.get("nav") is not None]
    navs = [r["nav"] for r in nav_rows if r.get("nav") is not None]
    top_dd = compute_top_drawdowns(dates, navs, 3)

    # Excel「指标汇总」中的交易次数可能与明细行数不一致，以明细为准
    summary_part["metrics"]["trade_count"] = len(trades)

    # Excel 的 Alpha/Beta 公式在部分周期（如 30 年跨度）可能失败或填 0，
    # 仅在 Excel 值同为 0（哨兵）时才用 nav + ^IXIC 的 CAPM 回归补算；
    # 已有非零的周期保留 Excel 口径。
    ex_a = summary_part["metrics"].get("alpha_pct")
    ex_b = summary_part["metrics"].get("beta")
    if (ex_a in (None, 0, 0.0)) and (ex_b in (None, 0, 0.0)):
        try:
            cagr_pct = summary_part["metrics"].get("cagr_pct")
            cagr_ann = cagr_pct / 100.0 if isinstance(cagr_pct, (int, float)) else None
            alpha_pct, beta = compute_alpha_beta_for_nav(nav_rows, cagr_ann)
            if alpha_pct is not None and beta is not None:
                summary_part["metrics"]["alpha_pct"] = alpha_pct
                summary_part["metrics"]["beta"] = beta
        except Exception as e:  # noqa: BLE001 - 网络失败时保留 Excel 原值
            print(f"[warn] {period}: alpha/beta 补算失败，保留 Excel 原值：{e}", file=sys.stderr)

    summary_obj: dict[str, Any] = {
        "version": version,
        "period": period,
        "symbols": summary_part["meta"]["symbols"],
        "start_date": summary_part["meta"]["start_date"],
        "end_date": summary_part["meta"]["end_date"],
        "initial_capital": summary_part["meta"]["initial_capital"],
        "commission_rate": summary_part["meta"]["commission_rate"],
        "slippage_bps": summary_part["meta"]["slippage_bps"],
        "metrics": summary_part["metrics"],
        "top_drawdowns": top_dd,
        "nav_rows": len(nav_rows),
        "trade_rows": len(trades),
    }

    stem = f"{version}-{period}"
    if dry_run:
        print(f"[dry-run] {stem}: nav={len(nav_rows)} trades={len(trades)} top1_dd={top_dd[0] if top_dd else None}")
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / f"{stem}-summary.json").write_text(
        json.dumps(summary_obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (out_dir / f"{stem}-nav.json").write_text(
        json.dumps(nav_rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (out_dir / f"{stem}-trades.json").write_text(
        json.dumps(trades, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Wrote {stem}-*.json ({len(nav_rows)} nav, {len(trades)} trades)")


def main() -> None:
    ap = argparse.ArgumentParser(description="Import backtest Excel to JSON")
    ap.add_argument("--10y", dest="y10", type=Path, help="10年回测 xlsx")
    ap.add_argument("--20y", dest="y20", type=Path, help="20年回测 xlsx")
    ap.add_argument("--30y", dest="y30", type=Path, help="30年回测 xlsx")
    ap.add_argument(
        "--out-dir",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "data" / "backtest",
        help="输出目录（默认 data/backtest）",
    )
    ap.add_argument("--version", default="v1.3.1", help="版本前缀")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--recompute-risk",
        action="store_true",
        help="不解析 Excel，仅基于现有 nav.json + ^IXIC 重算 summary.json 中的 alpha_pct / beta",
    )
    ap.add_argument(
        "--enrich-benchmark",
        action="store_true",
        help="不解析 Excel，为 nav.json 写入 port_ret_pct / qqq_bh_pct / qqq_dca_pct，并更新 summary.benchmark",
    )
    ap.add_argument(
        "--force",
        action="store_true",
        help="与 --recompute-risk / --enrich-benchmark 搭配：强制覆盖",
    )
    args = ap.parse_args()

    if args.recompute_risk:
        periods = ["10y", "20y", "30y"]
        recompute_summary_risk(args.out_dir, args.version, periods, force=args.force)
        return

    if args.enrich_benchmark:
        periods = ["10y", "20y", "30y"]
        enrich_benchmark(args.out_dir, args.version, periods, force=args.force)
        return

    mapping: list[tuple[str, Path | None]] = [
        ("10y", args.y10),
        ("20y", args.y20),
        ("30y", args.y30),
    ]
    any_set = any(p for _, p in mapping if p)
    if not any_set:
        ap.error(
            "请至少指定 --10y / --20y / --30y 之一，或 --recompute-risk / --enrich-benchmark"
        )
    for period, p in mapping:
        if p is None:
            continue
        if not p.is_file():
            print(f"文件不存在: {p}", file=sys.stderr)
            sys.exit(1)
        process_one_workbook(p, period, args.version, args.dry_run, args.out_dir)


if __name__ == "__main__":
    main()
